# ESSE É UM CÓDIGO ESCRITO PARA PEGAR UMA PASTA DO PC E LER OS ARQUIVOS PDFs NELA.
# DENTRO DESSES PDFS VOU TIRAR INFORMAÇÕES ESPECIFICAS
# COLOCAR ESSAS INFORMAÇÕES EM UM ARQUIVO EXCEL NA ABA "Valores_PDF"

import pdfplumber
import re
from pathlib import Path
import openpyxl

# Função para obter os nomes dos arquivos da pasta
def listando_arquivos(pasta):
    return [arquivo.name for arquivo in Path(pasta).iterdir() if arquivo.is_file()]

# Função para ajeitar o numero que vem no PDF para um valor numérico
def limpar_numero(valor):
    #Converte número contábil de string brasileira para float (ex: '1.234,56' → 1234.56)
    if valor:
        valor = valor.replace(".", "").replace(",", ".")  # Remove separador de milhar e ajusta decimal
        try:
            return float(valor)
        except ValueError:
            return None
    return None

# Função para puxar os valores de um PDF
def valores_PDFreembolso_excel(caminho_pasta, caminho_excel):

    try: # O "try" é para tentar rodar a função, caso falhe, ele vai para o "except" e mostra o erro
        # Tenta abrir o arquivo existente
        wb = openpyxl.load_workbook(caminho_excel)

        # Seleciona a aba "Valores_PDF"
        if "Valores_PDF" in wb.sheetnames:
            aba = wb["Valores_PDF"]
            # Apaga a aba (pra limpar os dados)
            wb.remove(aba)
            # Cria uma nova
            aba = wb.create_sheet("Valores_PDF")
        else:
            # Cria uma aba
            aba = wb.create_sheet("Valores_PDF")

        # Cria o cabeçalho
        aba.append(["ARQUIVO", "valor_1", "valor_2","OBS"])

        # Lista de arquivos na pasta
        filenames = listando_arquivos(caminho_pasta)

        # Processamento de cada arquivo
        for arquivo in filenames:

            try: 
                # Caminho completo do arquivo
                caminho_arquivo = caminho_pasta / arquivo

                # Inicializa as variáveis como None a cada iteração
                valor_1 = None
                valor_2 = None

                # Abre o PDF e extrai o texto da primeira página
                with pdfplumber.open(caminho_arquivo) as pdf:
                    texto = pdf.pages[0].extract_text() # MUDE AQUI O NÚMERO DA PÁGINA SE FOR NECESSÁRIO

                    # VALIDAÇÃO DE PÁGINA COM TEXTO
                    # Se a extração de texto falhar, envia mensagem de erro e pula para o próximo PDF
                    if texto is None:
                        print(f"Extração do arquivo {arquivo} falhou") # Mostra no terminal
                        aba.append([arquivo, "ERRO: A extraçao falhou"]) # Adiciona linha no Excel
                        continue

                    # VALIDAÇÃO DE TÍTULO DENTRO DO PDF
                    # Procura pelas informações específicas
                    # Checa se o o PDF ta com o título padrão dentro do PDF, se falhar envia mensagem de erro e pula para o próximo PDF
                    if "Relatório Mensal de Recarga" not in texto:
                        print(f"Arquivo {arquivo} não foi identificado como Relatório de Reembolso") # Mostra no terminal
                        print("-" * 50)
                        aba.append([arquivo, None, None, "ERRO: Não é um Relatório de Reembolso"]) # Adiciona linha no Excel
                        continue

                    # VALIDAÇÃO DE PERÍODO DENTRO DO PDF
                    # Agora vamos fazer uma checagem períodica para ver se o PDF está no mês correto (usamos os "-----" para sinalizar a troca mensal)
# -------------------------------------- MUDAR TODO MÊS ------------------------------------------------------------------------------------------------------
                    # Checa o período do PDF, se falhar envia mensagem de erro e pula para o próximo PDF
                    if "Período: 01/03/2025 até 31/03/2025" not in texto:
                        print(f"Arquivo {arquivo} está com o período de reembolso errado") # Mostra no terminal
                        print("-" * 50)
                        aba.append([arquivo, None, None, "ERRO: Está com o período de reembolso errado"]) # Adiciona linha no Excel
                        continue

                    # Divide o texto do PDF em linhas como se fosse uma tabela
                    linhas = texto.split("\n")                    
# ------------------------------------------------------------------------------------------------------------------------------------------------------------

                    # VALIDAÇÃO DE NOME DO ARQUIVO UTILIZANDO UM NOME PADRÃO A TODOS
                    # Agora vamos usar o regex para checar se o nome do arquivo está seguindo o padrão
                    # Caso não esteja, envia mensagem de erro e pula para o próximo PDF
                    match = re.search(r"Reembolso\s+(.*?)\s+- Mar 25", arquivo) # o padrão começa com "Reembolso", tem um meio variável e termina com " - Mar 25"
                    if match:
                        meio_texto = match.group(1) # Puxa só o meio do texto (Que é variável)
                    else:
                        print(f"Arquivo: {arquivo} não ta nomeado certo") # Mostra no terminal
                        print("-" * 50)
                        aba.append([arquivo, None, None, f"ERRO: Arquivo {arquivo} não ta nomeado certo"]) # Adiciona linha no Excel
                        continue

                    # VALIDAÇÃO DE NOME DO ARQUIVO DENTRO DO PDF
                    # Verifica se o arquivo está nomeado errado (o meio do nome não está na linha do PDF que tem o nome)
                    if meio_texto not in linhas[4]:
                        print(f"Arquivo: {arquivo} não ta nomeado certo") # Mostra no terminal
                        print("-" * 50)
                        aba.append([meio_texto, None, None, "ERRO: O nome do arquivo não condiz com o nome dentro do PDF"]) # Adiciona linha no Excel
                        continue

                    for linha in linhas:
                        # Captura o Valor 1, sendo puxado pela linha dentro do PDF que contém "Valor 1:"
                        if "Valor 1:" in linha:
                            match = re.search(r"R\$\s*([\d,.]+)", linha) # Puxa o valor que vem depois do "R$" (pois queremos apenas o número para fazer cálculos)
                            if match:
                                valor_1 = limpar_numero(match.group(1)) # Usa a função para ajeitar o número
                            else:
                                print(f"Valor 1 não encontrado no arquivo: {arquivo}") # Caso não consiga achar, mostra no terminal

                        # Captura o Valor 2 (No nosso caso é um valor de energia), sendo puxado pela linha dentro do PDF que contém "Valor 2:"
                        elif "Valor 2: " in linha:
                            match = re.search(r"([\d,.]+)\s*kWh", linha) # Puxa o valor que vem antes do " kWh" (pois queremos apenas o número para fazer cálculos)
                            if match:
                                valor_2 = limpar_numero(match.group(1))
                            else:
                                print(f"Consumo total kWh não encontrado no arquivo: {arquivo}") # Caso não consiga achar, mostra no terminal
                    
                    # Caso o valor 1 (tarifa) = R$ 2,07, exibe mensagem indicando erro (Tarifa errada no sistema), zera o valor 2 e pula para o próximo PDF
                    if valor_1 == 2.07:
                        print(f"Valores extraídos do arquivo: {arquivo} (TARIFA 2,07 ERRADA)") # Mostra no terminal
                        print("-" * 50)
                        aba.append([meio_texto, valor_1, 0, "ERRO: Tarifa no sistema de 2,07"]) # Adiciona linha no Excel
                        continue

                    # Caso o valor 2 = 0, exibe mensagem indicando erro (Não pode ser 0) e pula para o próximo PDF
                    if valor_2 == 0:
                        print(f"Valores extraídos do arquivo: {arquivo} (ERRO NA ENERGIA)") # Mostra no terminal
                        print("-" * 50)
                        aba.append([meio_texto, valor_1,valor_2]) # Adiciona linha no Excel
                        continue

                    # Dando tudo certo, exibe os valores extraídos
                    print(f"Valores extraídos do arquivo: {arquivo}") # Mostra no terminal
                    print("-" * 50)
                    aba.append([meio_texto, valor_1,valor_2]) # Adiciona linha no Excel
            
            # Trata os erros indicando o arquivo problemático e continua o processamento dos demais arquivos
            except Exception as e:    
                print(f"Erro ao processar {arquivo}: {str(e)}") # Mostra no terminal
                aba.append([arquivo, None, None, f"ERRO: {str(e)}"]) # Adiciona linha no Excel
                
        
        # Salva o Excel
        wb.save(caminho_excel)

        # A função retorna uma mensagem se conclusão
        return f"Salvamento concluído e salvo na {aba}" # Mostra no terminal
    # Aqui mostra erros fatais, que impediram a abertura da pasta ou do arquivo Excel
    except Exception as e:
        return f"Erro fatal: {str(e)}" # Mostra no terminal

# Diretório dos PDFs
caminho_pasta = Path(r"INSIRA AQUI O CAMINHO DA PASTA DOS PDFS") # Mude aqui o caminho da pasta

# Caminho do arquivo Excel
caminho_excel = r"INSIRA AQUI O CAMINHO DO ARQUIVO EXCEL" # Mude aqui o caminho do arquivo Excel

print(valores_PDFreembolso_excel(caminho_pasta,caminho_excel))
